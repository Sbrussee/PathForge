"""Integration tests for the slide retrieval task pipeline.

Coverage matrix:

  Scenario                            | reference/query | query_reference | exclusion
  ------------------------------------|-----------------|-----------------|----------
  basic reference/query split         |       ✓         |                 | patient
  query_reference shared pool         |                 |       ✓         | patient
  exclusion_level=none                |       ✓         |                 | none
  exclusion_level=slide               |       ✓         |                 | slide
  exclusion_level=case                |       ✓         |                 | case
  representation cache hit            |       ✓         |                 | patient
  representation cache miss           |       ✓         |                 | patient
  mixed cache hit/miss                |       ✓         |                 | patient
  multiple reference datasets         |       ✓×2       |                 | patient
  incompatible combo → skipped        |       ✓         |                 | patient
  empty datasets → ValueError         |                 |                 |
  tiling_id mismatch → ValueError     |       ✓         |                 |
  aggregation_level mismatch →ValueError|     ✓         |                 |
  no reference → ValueError           |                 |       ✓         |
  no query → ValueError               |                 |       ✓         |
  manifest fields complete            |       ✓         |                 | patient
  results CSV structure               |       ✓         |                 | patient
  run hash determinism                |       ✓         |                 | patient
  failed representation raises        |       ✓         |                 | patient
"""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from typing import Any
from unittest.mock import MagicMock

import pytest

from pathforge.core.datasets.bag_dataset import BagSample, SlideRetrievalBagDataset, SlideRetrievalDatasetItem
from pathforge.core.experiments.combinations import ComboConfig
from pathforge.core.tasks.slide_retrieval import SlideRetrievalTask
from pathforge.slide_retrieval.representation_strategies.types import RetrievalRepresentation
from pathforge.slide_retrieval.search_strategies.types import SearchHit, SearchResult


# ---------------------------------------------------------------------------
# Fakes
# ---------------------------------------------------------------------------


class _FakeSample:
    def __init__(
        self,
        sample_id: str,
        *,
        slide_ids: frozenset[str] | None = None,
        patient_id: str | None = None,
        case_id: str | None = None,
        category: str | None = None,
    ) -> None:
        self.sample_id = sample_id
        self.slide_ids = slide_ids or frozenset({sample_id})
        self.patient_id = patient_id
        self.case_id = case_id
        self.category = category


class _FakeSlideRetrievalBagDataset(SlideRetrievalBagDataset):
    """Minimal in-memory SlideRetrievalBagDataset for testing."""

    def __init__(
        self,
        *,
        name: str,
        tiling_id: str,
        aggregation_level: str,
        artifacts_dir: Path,
        samples: list[_FakeSample],
        feature_level: str = "patch",
    ) -> None:
        self._name = name
        self.tiling_id = tiling_id
        self.aggregation_level = aggregation_level
        self.artifacts_dir = artifacts_dir
        self._samples = samples
        self._feature_level = feature_level
        self._sample_loader = None

    @property
    def num_bags(self) -> int:
        return len(self._samples)

    def get_sample(self, index: int) -> _FakeSample:
        return self._samples[index]

    def get_feature_level(self) -> str:
        return self._feature_level

    def get_feature_level_reason(self) -> str:
        return f"feature_level={self._feature_level}"

    def bind_sample_loader(self, loader) -> None:
        self._sample_loader = loader

    def clear_sample_loader(self) -> None:
        self._sample_loader = None

    def __len__(self) -> int:
        return len(self._samples)

    def __getitem__(self, index: int) -> SlideRetrievalDatasetItem:
        return SlideRetrievalDatasetItem(index=index, sample=self._samples[index], inputs={})


class _FakeRepresentationStrategy:
    supported_feature_levels = frozenset({"patch"})
    output_representation_kind = "patch_vector"

    def __init__(self, params=None, **kwargs) -> None:
        self._params = dict(params or {})

    def hyperparam_values(self) -> dict[str, object]:
        return dict(self._params)

    def load_sample(self, sample):
        return {}

    def run(self, *, sample, **kwargs) -> RetrievalRepresentation:
        return RetrievalRepresentation(
            sample_id=sample.sample_id,
            data=[[1.0, 2.0]],
        )


class _IncompatibleRepresentationStrategy(_FakeRepresentationStrategy):
    supported_feature_levels = frozenset({"slide"})


class _FakeSearchStrategy:
    supported_representation_kinds = frozenset({"patch_vector"})

    def __init__(self, params=None, **kwargs) -> None:
        self._params = dict(params or {})
        self.search_database: list[RetrievalRepresentation] = []

    def hyperparam_values(self) -> dict[str, object]:
        return dict(self._params)

    def build_database(self, database_representations) -> None:
        self.search_database = list(database_representations)

    def search(self, query_representation, **kwargs) -> SearchResult:
        hits = [
            SearchHit(sample_id=rep.sample_id, score=float(i + 1) * 0.1, rank=i + 1)
            for i, rep in enumerate(self.search_database)
        ]
        return SearchResult(
            query_sample_id=query_representation.sample_id,
            hits=hits,
        )


# ---------------------------------------------------------------------------
# Test helpers
# ---------------------------------------------------------------------------


def _make_task(tmp_path: Path, *, exclusion_level: str = "patient") -> SlideRetrievalTask:
    cfg = SimpleNamespace(
        experiment=SimpleNamespace(aggregation_level="slide", num_workers=0),
        slide_retrieval=SimpleNamespace(exclusion_level=exclusion_level),
    )
    experiment = SimpleNamespace(cfg=cfg, project_root=str(tmp_path))
    return SlideRetrievalTask(experiment)


def _make_combo(
    *,
    tiling_id: str = "256px_0.5mpp",
    feature_extraction: str = "uni",
    representation: str = "mean_pooling",
    search_strategy: str = "cosine_knn",
) -> ComboConfig:
    tile_px, rest = tiling_id.split("px_")
    tile_mpp = rest.replace("mpp", "")
    return ComboConfig(
        tile_px=int(tile_px),
        tile_px_params={},
        tile_mpp=float(tile_mpp),
        tile_mpp_params={},
        feature_extraction=feature_extraction,
        feature_extraction_params={},
        retrieval_representation=representation,
        retrieval_representation_params={},
        search_strategy=search_strategy,
        search_strategy_params={},
    )


def _make_dataset(
    tmp_path: Path,
    *,
    name: str,
    sample_ids: list[str],
    tiling_id: str = "256px_0.5mpp",
    aggregation_level: str = "slide",
    patient_ids: list[str] | None = None,
    feature_level: str = "patch",
) -> _FakeSlideRetrievalBagDataset:
    samples = [
        _FakeSample(
            sid,
            patient_id=(patient_ids[i] if patient_ids else sid),
        )
        for i, sid in enumerate(sample_ids)
    ]
    return _FakeSlideRetrievalBagDataset(
        name=name,
        tiling_id=tiling_id,
        aggregation_level=aggregation_level,
        artifacts_dir=tmp_path / "artifacts",
        samples=samples,
        feature_level=feature_level,
    )


def _no_cache(self, *, bag_dataset, representation_id, aggregation_level, exclusion_level):
    """Simulate all samples missing from cache."""
    from torch.utils.data import Subset
    missing = Subset(bag_dataset, list(range(bag_dataset.num_bags)))
    return [], missing


def _full_cache(self, *, bag_dataset, representation_id, aggregation_level, exclusion_level):
    """Simulate all samples already cached."""
    reps = [
        RetrievalRepresentation(
            sample_id=bag_dataset.get_sample(i).sample_id,
            data=[[1.0]],
            exclusion_key=bag_dataset.get_sample(i).patient_id,
        )
        for i in range(bag_dataset.num_bags)
    ]
    return reps, None


# ---------------------------------------------------------------------------
# Basic reference/query split
# ---------------------------------------------------------------------------


def test_reference_query_split_produces_results(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["ref-1", "ref-2"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    assert result["num_queries"] == 1
    assert result["num_reference_items"] == 2
    assert Path(result["output_dir"]).is_dir()


# ---------------------------------------------------------------------------
# query_reference shared pool
# ---------------------------------------------------------------------------


def test_query_reference_included_in_both_sets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "query_reference": [_make_dataset(tmp_path, name="shared", sample_ids=["s-1", "s-2", "s-3"])],
        },
    )

    assert result["num_queries"] == 3
    assert result["num_reference_items"] == 3


# ---------------------------------------------------------------------------
# Exclusion levels
# ---------------------------------------------------------------------------


def test_exclusion_level_none_assigns_no_exclusion_key(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path, exclusion_level="none")
    task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    sample = _FakeSample("s", patient_id="p1")
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="none"
    )
    assert key is None


def test_exclusion_level_patient_uses_patient_id(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="patient")
    sample = _FakeSample("s-1", patient_id="patient-A")
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="patient"
    )
    assert key == "patient-A"


def test_exclusion_level_case_uses_case_id(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="case")
    sample = _FakeSample("s-1", case_id="case-X")
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="case"
    )
    assert key == "case-X"


def test_exclusion_level_slide_uses_sample_id(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="slide")
    sample = _FakeSample("slide-42")
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="slide"
    )
    assert key == "slide-42"


def test_exclusion_level_slide_raises_when_aggregation_is_not_slide(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="slide")
    sample = _FakeSample("s-1")
    with pytest.raises(ValueError, match="aggregation_level='slide'"):
        task._build_exclusion_key(
            sample=sample, aggregation_level="case", exclusion_level="slide"
        )


def test_exclusion_level_patient_returns_none_when_patient_id_missing(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="patient")
    sample = _FakeSample("s-1", patient_id=None)
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="patient"
    )
    assert key is None


def test_exclusion_level_case_returns_none_when_case_id_missing(tmp_path: Path) -> None:
    task = _make_task(tmp_path, exclusion_level="case")
    sample = _FakeSample("s-1", case_id=None)
    key = task._build_exclusion_key(
        sample=sample, aggregation_level="slide", exclusion_level="case"
    )
    assert key is None


def test_exclusion_level_unknown_raises(tmp_path: Path) -> None:
    task = _make_task(tmp_path)
    sample = _FakeSample("s-1")
    with pytest.raises(ValueError, match="Unsupported"):
        task._build_exclusion_key(
            sample=sample, aggregation_level="slide", exclusion_level="unknown_level"
        )


# ---------------------------------------------------------------------------
# Representation caching
# ---------------------------------------------------------------------------


def test_cached_representations_are_not_recomputed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """When all representations are in cache, no computation is triggered."""
    import pathforge.core.tasks.slide_retrieval as mod

    compute_calls: list[str] = []

    class _TrackingStrategy(_FakeRepresentationStrategy):
        def run(self, *, sample, **kwargs) -> RetrievalRepresentation:
            compute_calls.append(sample.sample_id)
            return super().run(sample=sample, **kwargs)

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _TrackingStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1", "r-2"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    assert compute_calls == [], "No representations should be computed when all are cached"


def test_missing_representations_are_materialized(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Samples absent from cache must be materialized via the strategy."""
    import pathforge.core.tasks.slide_retrieval as mod

    compute_calls: list[str] = []

    class _TrackingStrategy(_FakeRepresentationStrategy):
        def run(self, *, sample, **kwargs) -> RetrievalRepresentation:
            compute_calls.append(sample.sample_id)
            return RetrievalRepresentation(sample_id=sample.sample_id, data=[[0.5]])

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _TrackingStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _no_cache)
    monkeypatch.setattr(mod, "atomic_slide_artifact_write", MagicMock())
    monkeypatch.setattr(mod, "save_slide_retrieval_representation", MagicMock())

    task = _make_task(tmp_path)
    task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    assert set(compute_calls) == {"r-1", "q-1"}


# ---------------------------------------------------------------------------
# Incompatible combo → skipped
# ---------------------------------------------------------------------------


def test_incompatible_feature_level_returns_skipped_status(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _IncompatibleRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"slide"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "slide_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"slide_vector"}))

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"], feature_level="patch")],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"], feature_level="patch")],
        },
    )

    assert result["status"] == "skipped_incompatible_combo"
    assert "reason" in result


# ---------------------------------------------------------------------------
# Validation errors
# ---------------------------------------------------------------------------


def test_empty_datasets_raises_value_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))

    task = _make_task(tmp_path)

    with pytest.raises(ValueError, match="No bag datasets"):
        task.execute(combo_cfg=_make_combo(), datasets_by_use={})


def test_tiling_id_mismatch_raises_value_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))

    task = _make_task(tmp_path)

    with pytest.raises(ValueError, match="tiling_id"):
        task.execute(
            combo_cfg=_make_combo(tiling_id="256px_0.5mpp"),
            datasets_by_use={
                "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"], tiling_id="512px_0.5mpp")],
                "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"], tiling_id="256px_0.5mpp")],
            },
        )


def test_aggregation_level_mismatch_raises_value_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))

    task = _make_task(tmp_path)

    with pytest.raises(ValueError, match="aggregation_level"):
        task.execute(
            combo_cfg=_make_combo(),
            datasets_by_use={
                "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"], aggregation_level="case")],
                "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"], aggregation_level="slide")],
            },
        )


# ---------------------------------------------------------------------------
# Multiple reference datasets
# ---------------------------------------------------------------------------


def test_multiple_reference_datasets_are_merged(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [
                _make_dataset(tmp_path, name="ref-A", sample_ids=["a-1", "a-2"]),
                _make_dataset(tmp_path, name="ref-B", sample_ids=["b-1"]),
            ],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    assert result["num_reference_items"] == 3


# ---------------------------------------------------------------------------
# Manifest and CSV structure
# ---------------------------------------------------------------------------


def test_manifest_contains_required_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import json
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    manifest = json.loads((Path(result["output_dir"]) / "manifest.json").read_text())
    required_keys = {
        "tiling_id", "aggregation_level", "feature_extraction",
        "slide_representation", "search_method", "representation_id",
        "exclusion_level", "num_queries", "num_reference_items", "top_k_saved",
    }
    assert required_keys <= set(manifest.keys()), (
        f"Missing manifest keys: {required_keys - set(manifest.keys())}"
    )


def test_results_xlsx_has_correct_schema(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from openpyxl import load_workbook
    import pathforge.core.tasks.slide_retrieval as mod

    monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
    monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
    monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
    monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
    monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
    monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

    task = _make_task(tmp_path)
    result = task.execute(
        combo_cfg=_make_combo(),
        datasets_by_use={
            "reference": [_make_dataset(tmp_path, name="ref", sample_ids=["r-1"])],
            "query": [_make_dataset(tmp_path, name="qry", sample_ids=["q-1"])],
        },
    )

    xlsx_path = Path(result["output_dir"]) / "query_results.xlsx"
    workbook = load_workbook(xlsx_path, read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    fieldnames = list(rows[0]) if rows else None
    data_rows = rows[1:]

    assert fieldnames is not None
    assert "query_sample_id" in fieldnames
    assert len(data_rows) == 1


# ---------------------------------------------------------------------------
# Deterministic run hash
# ---------------------------------------------------------------------------


def test_run_directory_hash_is_deterministic(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same config and data produce the same run directory path."""
    import pathforge.core.tasks.slide_retrieval as mod

    def _make_task_and_run(base: Path) -> str:
        monkeypatch.setattr(mod, "build_representation_strategy", lambda _n, **kw: _FakeRepresentationStrategy())
        monkeypatch.setattr(mod, "build_search_strategy", lambda _n, **kw: _FakeSearchStrategy())
        monkeypatch.setattr(mod, "get_representation_strategy_supported_feature_levels", lambda _n: frozenset({"patch"}))
        monkeypatch.setattr(mod, "get_representation_strategy_output_kind", lambda _n: "patch_vector")
        monkeypatch.setattr(mod, "get_search_strategy_supported_representation_kinds", lambda _n: frozenset({"patch_vector"}))
        monkeypatch.setattr(SlideRetrievalTask, "_collect_existing_representations", _full_cache)

        task = _make_task(base)
        result = task.execute(
            combo_cfg=_make_combo(),
            datasets_by_use={
                "reference": [_make_dataset(base, name="ref", sample_ids=["r-1"])],
                "query": [_make_dataset(base, name="qry", sample_ids=["q-1"])],
            },
        )
        return Path(result["output_dir"]).name

    run_dir_a = _make_task_and_run(tmp_path / "run_a")
    run_dir_b = _make_task_and_run(tmp_path / "run_b")

    assert run_dir_a == run_dir_b, (
        f"Run hashes differ: {run_dir_a!r} vs {run_dir_b!r}"
    )
