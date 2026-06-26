"""Extended end-to-end smoke tests for slide retrieval.

These tests exercise more complex scenarios than the basic benchmark smoke test:
- Multi-combo grid producing separate run directories
- query_reference round-trip
- Representation cache reuse across two executions
- Failed representation materialisation bubbling up as RuntimeError
- Unsupported dataset use raises ValueError
- No reference representations raises ValueError
- No query representations raises ValueError
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from pathforge.core.experiments.combinations import ComboConfig


def _read_query_results_xlsx(path: Path) -> tuple[list[str] | None, list[dict[str, object]]]:
    """Read a ranked query-results Excel workbook into (fieldnames, row dicts)."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return None, []
    fieldnames = [str(value) for value in rows[0]]
    records = [dict(zip(fieldnames, row)) for row in rows[1:]]
    return fieldnames, records
from pathforge.core.tasks.slide_retrieval import SlideRetrievalTask
from pathforge.slide_retrieval.representation_strategies.types import (
    RetrievalRepresentation,
)
from pathforge.slide_retrieval.search_strategies.types import SearchHit, SearchResult
from ._smoke_dataset import attach_smoke_outputs, capture_smoke_metrics
from .conftest import RetrievalDatasets


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_task(tmp_path: Path, *, exclusion_level: str = "patient") -> SlideRetrievalTask:
    from types import SimpleNamespace

    cfg = SimpleNamespace(
        experiment=SimpleNamespace(aggregation_level="slide", num_workers=0),
        slide_retrieval=SimpleNamespace(exclusion_level=exclusion_level),
    )
    return SlideRetrievalTask(SimpleNamespace(cfg=cfg, project_root=str(tmp_path)))


def _make_combo(
    *, representation: str = "hshr-features", search: str = "yottixel"
) -> ComboConfig:
    return ComboConfig(
        tile_px=224,
        tile_px_params={},
        tile_mpp=1.0,
        tile_mpp_params={},
        feature_extraction="resnet18",
        feature_extraction_params={},
        retrieval_representation=representation,
        retrieval_representation_params={},
        search_strategy=search,
        search_strategy_params={},
    )


def _full_cache(self, *, bag_dataset, representation_id, aggregation_level, exclusion_level):
    """Return pre-cached representations for all samples — bypasses H5 access."""
    reps = [
        RetrievalRepresentation(
            sample_id=bag_dataset.get_sample(i).sample_id,
            data=[[1.0]],
            exclusion_key=bag_dataset.get_sample(i).patient_id,
        )
        for i in range(bag_dataset.num_bags)
    ]
    return reps, None


def _register_retrieval_strategies() -> None:
    """Register all production representation and search strategies."""
    from pathforge.slide_retrieval.representation_strategies.registry import (
        import_representation_strategy_modules,
    )
    from pathforge.slide_retrieval.search_strategies.registry import (
        import_search_strategy_modules,
    )

    import_representation_strategy_modules()
    import_search_strategy_modules()


# ---------------------------------------------------------------------------
# Test-instrumentation strategy stubs (only for error / behaviour scenarios)
# ---------------------------------------------------------------------------


class _CountingRepresentationStrategy:
    """Counts run() invocations to verify cache reuse."""

    supported_feature_levels = frozenset({"patch"})
    output_representation_kind = "patch_vector"

    def __init__(self, log: list[str]) -> None:
        self._log = log

    def hyperparam_values(self) -> dict:
        return {}

    def load_sample(self, *, index=None, sample=None, base_dataset=None):
        return {}

    def run(self, *, sample, **kwargs) -> RetrievalRepresentation:
        self._log.append(sample.sample_id)
        return RetrievalRepresentation(sample_id=sample.sample_id, data=[[0.1]])


class _CrashingRepresentationStrategy:
    supported_feature_levels = frozenset({"patch"})
    output_representation_kind = "patch_vector"

    def hyperparam_values(self) -> dict:
        return {}

    def load_sample(self, *, index=None, sample=None, base_dataset=None):
        return {}

    def run(self, *, sample, **kwargs):
        raise RuntimeError("Simulated strategy crash")


class _EmptyRepresentationStrategy:
    supported_feature_levels = frozenset({"patch"})
    output_representation_kind = "patch_vector"

    def hyperparam_values(self) -> dict:
        return {}

    def load_sample(self, *, index=None, sample=None, base_dataset=None):
        return {}

    def run(self, *, sample, **kwargs):
        return RetrievalRepresentation(sample_id=sample.sample_id, data=[[0.0]])


# ---------------------------------------------------------------------------
# Multi-combo grid
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_multi_combo_grid_produces_separate_run_dirs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
    slide_level_feature_matrix: tuple,
) -> None:
    """Each (representation, search_strategy) combination writes a separate run dir."""
    import pathforge.core.tasks.slide_retrieval as mod

    _register_retrieval_strategies()
    slide_ids, feature_matrix = slide_level_feature_matrix

    def _real_feature_cache(
        self,
        *,
        bag_dataset,
        representation_id,
        aggregation_level,
        exclusion_level,
    ):
        reps = []
        for i in range(bag_dataset.num_bags):
            sample = bag_dataset.get_sample(i)
            try:
                idx = slide_ids.index(sample.sample_id)
                data = [feature_matrix[idx].tolist()]
            except ValueError:
                data = [feature_matrix[0].tolist()]
            reps.append(
                RetrievalRepresentation(
                    sample_id=sample.sample_id,
                    data=data,
                    exclusion_key=sample.patient_id,
                )
            )
        return reps, None

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _real_feature_cache
    )

    datasets = {
        "reference": [retrieval_wsi_datasets.reference],
        "query": [retrieval_wsi_datasets.query],
    }

    combos = [
        _make_combo(representation="hshr-features", search="yottixel"),
        _make_combo(representation="sdm-features", search="yottixel"),
        _make_combo(representation="hshr-features", search="retccl"),
    ]

    task = _make_task(tmp_path)
    output_dirs: set[str] = set()

    with capture_smoke_metrics(
        tmp_path / "metrics",
        step_name="smoke_retrieval_multi_combo_grid",
        metadata={"num_combos": len(combos)},
    ) as metadata:
        for combo in combos:
            result = task.execute(combo_cfg=combo, datasets_by_use=datasets)
            output_dirs.add(result["output_dir"])
        attach_smoke_outputs(
            metadata,
            step_name="smoke_retrieval_multi_combo_grid",
            final={
                f"run_dir_{i}": Path(d)
                for i, d in enumerate(sorted(output_dirs))
            },
        )

    assert len(output_dirs) == len(combos), (
        f"Expected {len(combos)} distinct run dirs, got {len(output_dirs)}: {output_dirs}"
    )


# ---------------------------------------------------------------------------
# query_reference round-trip
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_query_reference_self_retrieval(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
    slide_level_feature_matrix: tuple,
) -> None:
    """Real Yottixel self-retrieval over a query_reference pool.

    Each slide is both a query and a reference, backed by its *own* distinct
    real GTEx feature vector. With ``exclusion_level='none'`` (so nothing is
    filtered out), a correct search must return every query as its own rank-1
    hit at distance 0 — this verifies the search actually ranks by similarity,
    not just that the pipeline emits well-formed rows.
    """
    _register_retrieval_strategies()
    slide_ids, feature_matrix = slide_level_feature_matrix

    def _distinct_real_cache(
        self, *, bag_dataset, representation_id, aggregation_level, exclusion_level
    ):
        """Return a distinct real per-slide representation with no exclusion key."""
        reps = []
        for i in range(bag_dataset.num_bags):
            sample = bag_dataset.get_sample(i)
            idx = slide_ids.index(sample.sample_id)  # pool slides are a subset of slide_ids
            reps.append(
                RetrievalRepresentation(
                    sample_id=sample.sample_id,
                    data=[feature_matrix[idx].tolist()],
                    exclusion_key=None,  # no exclusion -> a query may retrieve itself
                )
            )
        return reps, None

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _distinct_real_cache
    )

    pool_dataset = retrieval_wsi_datasets.reference  # reuse 10-slide dataset as pool
    pool_ids = {
        pool_dataset.get_sample(i).sample_id for i in range(pool_dataset.num_bags)
    }
    task = _make_task(tmp_path, exclusion_level="none")

    with capture_smoke_metrics(
        tmp_path / "metrics",
        step_name="smoke_retrieval_self_retrieval",
        metadata={"pool_size": pool_dataset.num_bags, "exclusion_level": "none"},
    ) as metadata:
        result = task.execute(
            combo_cfg=_make_combo(representation="hshr-features", search="yottixel"),
            datasets_by_use={"query_reference": [pool_dataset]},
        )
        output_dir = Path(result["output_dir"])
        attach_smoke_outputs(
            metadata,
            step_name="smoke_retrieval_self_retrieval",
            final={"query_results_xlsx": output_dir / "query_results.xlsx"},
        )

    assert result["num_queries"] == pool_dataset.num_bags
    assert result["num_reference_items"] == pool_dataset.num_bags

    fieldnames, rows = _read_query_results_xlsx(output_dir / "query_results.xlsx")
    assert fieldnames is not None and "rank_1_sample_id" in fieldnames
    assert len(rows) == pool_dataset.num_bags

    rank_id_cols = sorted(
        (int(name.split("_")[1]), name)
        for name in fieldnames
        if name.startswith("rank_") and name.endswith("_sample_id")
    )

    for row in rows:
        query_id = row["query_sample_id"]

        # Real correctness: the query retrieves *itself* as rank-1 at distance 0.
        assert row["rank_1_sample_id"] == query_id, (
            f"query {query_id} did not self-retrieve; rank-1 was "
            f"{row.get('rank_1_sample_id')!r}"
        )
        assert float(row["rank_1_score"]) == pytest.approx(0.0, abs=1e-6)

        # Ranking sanity: contiguous ranks, hits are real pool members,
        # scores monotonically non-decreasing (Yottixel score == distance).
        present_ranks = [r for r, col in rank_id_cols if row.get(col)]
        assert present_ranks == list(range(1, len(present_ranks) + 1))
        previous_score = None
        for rank in present_ranks:
            hit_id = row[f"rank_{rank}_sample_id"]
            assert hit_id in pool_ids
            score = float(row[f"rank_{rank}_score"])
            if previous_score is not None:
                assert score >= previous_score - 1e-9
            previous_score = score


# ---------------------------------------------------------------------------
# Representation cache reuse
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_representation_cache_reuse_across_executions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
) -> None:
    """Second execution with the same combo must reuse cached representations."""
    import pathforge.core.tasks.slide_retrieval as mod

    _register_retrieval_strategies()
    compute_log: list[str] = []
    counting_strategy = _CountingRepresentationStrategy(compute_log)

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _full_cache
    )
    monkeypatch.setattr(
        mod, "build_representation_strategy", lambda _n, **kw: counting_strategy
    )

    task = _make_task(tmp_path)
    datasets = {
        "reference": [retrieval_wsi_datasets.single_ref],
        "query": [retrieval_wsi_datasets.single_ref],
    }
    combo = _make_combo()

    task.execute(combo_cfg=combo, datasets_by_use=datasets)
    task.execute(combo_cfg=combo, datasets_by_use=datasets)

    assert compute_log == [], "Representations should be served from cache on both runs"


# ---------------------------------------------------------------------------
# Failed representation raises RuntimeError
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_failed_representation_materialisation_raises(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
) -> None:
    """A strategy crash during materialisation must surface as RuntimeError."""
    import pathforge.core.tasks.slide_retrieval as mod

    _register_retrieval_strategies()

    def _crash_cache(self, *, bag_dataset, representation_id, aggregation_level, exclusion_level):
        from torch.utils.data import Subset

        missing = Subset(bag_dataset, list(range(bag_dataset.num_bags)))
        return [], missing

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _crash_cache
    )
    monkeypatch.setattr(mod, "atomic_slide_artifact_write", MagicMock())
    monkeypatch.setattr(mod, "save_slide_retrieval_representation", MagicMock())
    monkeypatch.setattr(
        mod, "build_representation_strategy", lambda _n, **kw: _CrashingRepresentationStrategy()
    )

    task = _make_task(tmp_path)

    with pytest.raises(RuntimeError, match="representation creation failed"):
        task.execute(
            combo_cfg=_make_combo(),
            datasets_by_use={
                "reference": [retrieval_wsi_datasets.single_ref],
                "query": [retrieval_wsi_datasets.single_ref],
            },
        )


# ---------------------------------------------------------------------------
# Unsupported dataset use raises ValueError
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_unsupported_dataset_use_raises(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
) -> None:
    _register_retrieval_strategies()
    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _full_cache
    )

    task = _make_task(tmp_path)

    with pytest.raises(ValueError, match="Unsupported retrieval dataset use"):
        task.execute(
            combo_cfg=_make_combo(),
            datasets_by_use={"training": [retrieval_wsi_datasets.single_ref]},
        )


# ---------------------------------------------------------------------------
# No reference → ValueError
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_no_reference_representations_raises(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    retrieval_wsi_datasets: RetrievalDatasets,
) -> None:
    """When the reference set is empty after splitting, execution must raise."""
    import pathforge.core.tasks.slide_retrieval as mod

    _register_retrieval_strategies()

    def _empty_cache(self, *, bag_dataset, **kwargs):
        return [], None

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _empty_cache
    )
    monkeypatch.setattr(mod, "atomic_slide_artifact_write", MagicMock())
    monkeypatch.setattr(mod, "save_slide_retrieval_representation", MagicMock())
    monkeypatch.setattr(
        mod, "build_representation_strategy", lambda _n, **kw: _EmptyRepresentationStrategy()
    )

    task = _make_task(tmp_path)

    with pytest.raises(ValueError, match="No reference representations"):
        task.execute(
            combo_cfg=_make_combo(),
            datasets_by_use={"query": [retrieval_wsi_datasets.single_ref]},
        )


# ---------------------------------------------------------------------------
# Manifest and CSV correctness (real feature vectors)
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_manifest_and_csv_are_well_formed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    slide_level_feature_matrix: tuple,
    retrieval_wsi_datasets: RetrievalDatasets,
) -> None:
    import pathforge.core.tasks.slide_retrieval as mod

    _register_retrieval_strategies()

    # Use real GTEx feature vectors — one per slide from mean+max pooling.
    slide_ids, feature_matrix = slide_level_feature_matrix
    ref_dataset = retrieval_wsi_datasets.reference
    qry_dataset = retrieval_wsi_datasets.query

    def _real_feature_cache(
        self,
        *,
        bag_dataset,
        representation_id,
        aggregation_level,
        exclusion_level,
    ):
        reps = []
        for i in range(bag_dataset.num_bags):
            sample = bag_dataset.get_sample(i)
            try:
                idx = slide_ids.index(sample.sample_id)
                data = [feature_matrix[idx].tolist()]
            except ValueError:
                data = [feature_matrix[0].tolist()]
            reps.append(
                RetrievalRepresentation(
                    sample_id=sample.sample_id,
                    data=data,
                    exclusion_key=sample.patient_id,
                )
            )
        return reps, None

    monkeypatch.setattr(
        SlideRetrievalTask, "_collect_existing_representations", _real_feature_cache
    )

    task = _make_task(tmp_path)

    with capture_smoke_metrics(
        tmp_path / "metrics",
        step_name="smoke_retrieval_manifest_csv_well_formed",
        metadata={
            "num_queries": qry_dataset.num_bags,
            "num_reference": ref_dataset.num_bags,
            "representation": "hshr-features",
        },
    ) as metadata:
        result = task.execute(
            combo_cfg=_make_combo(representation="hshr-features", search="yottixel"),
            datasets_by_use={
                "reference": [ref_dataset],
                "query": [qry_dataset],
            },
        )
        output_dir = Path(result["output_dir"])
        attach_smoke_outputs(
            metadata,
            step_name="smoke_retrieval_manifest_csv_well_formed",
            final={
                "manifest_json": output_dir / "manifest.json",
                "query_results_xlsx": output_dir / "query_results.xlsx",
            },
        )

    manifest = json.loads((output_dir / "manifest.json").read_text())

    assert manifest["slide_representation"] == "hshr-features"
    assert manifest["search_method"] == "yottixel"
    assert manifest["num_queries"] == qry_dataset.num_bags
    assert manifest["num_reference_items"] == ref_dataset.num_bags
    assert manifest["exclusion_level"] == "patient"
    assert isinstance(manifest["representation_id"], str)
    assert len(manifest["representation_id"]) > 0

    fieldnames, rows = _read_query_results_xlsx(output_dir / "query_results.xlsx")

    assert fieldnames is not None
    assert "query_sample_id" in fieldnames
    assert len(rows) == qry_dataset.num_bags
    ref_ids = {ref_dataset.get_sample(i).sample_id for i in range(ref_dataset.num_bags)}
    qry_ids = {qry_dataset.get_sample(i).sample_id for i in range(qry_dataset.num_bags)}
    assert all(row["query_sample_id"] in qry_ids for row in rows)
